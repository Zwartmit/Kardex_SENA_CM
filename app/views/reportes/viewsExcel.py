import io
from io import BytesIO
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from app.models import Movimiento, DetalleMovimiento

@login_required
@never_cache
def exportar_excel(request):
    title_text = "Reporte de movimientos"
    headers = ["ID", "Descripción", "Elementos", "Cantidad", "Fecha", "Proyecto", "Ficha"]

    object_list = Movimiento.objects.all()
    data_rows = []

    for movimiento in object_list:
        elementos = DetalleMovimiento.objects.filter(movimiento=movimiento)

        for detalle in elementos:
            data_rows.append([
                movimiento.id,
                movimiento.descripcion,
                detalle.elemento.item,
                detalle.cantidad,
                movimiento.fecha.strftime("%d/%m/%Y"),
                movimiento.proyecto,
                movimiento.num_ficha
            ])

    filename = "Reporte de movimientos.xlsx"
    return generate_excel_report(title_text, headers, data_rows, filename)

def generate_excel_report(title_text, headers, data_rows, filename):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title_text

    worksheet.merge_cells('A3:G3')
    title_cell = worksheet.cell(row=3, column=1)
    title_cell.value = title_text
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.font = title_cell.font.copy(bold=True, size=14)

    border_style = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )

    for col in range(1, 8):
        cell = worksheet.cell(row=3, column=col)
        cell.border = border_style
        
    worksheet.append(headers)

    fill = PatternFill(start_color="04644B", end_color="04644B", fill_type="solid")
    for col in worksheet.iter_cols(min_row=4, max_row=4):
        for cell in col:
            cell.fill = fill
            cell.font = cell.font.copy(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in data_rows:
        worksheet.append(row)

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    for row in worksheet.iter_rows(min_row=4, max_row=len(data_rows) + 4, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_style

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response
