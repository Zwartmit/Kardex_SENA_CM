from flask import Flask, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
import io

app = Flask(__name__)

@app.route('/generate-excel', methods=['POST'])
def generate_excel():
    # Recibir datos del frontend (en formato JSON)
    data = request.get_json()

    # Crear el archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Formato SENA"

    # Estilo general
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Encabezados
    ws.merge_cells("A1:F1")
    ws["A1"] = "SISTEMA INTEGRADO DE GESTIÓN"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = alignment_center

    ws.merge_cells("A2:F2")
    ws["A2"] = "Formato: Entrega de materiales de formación al aprendiz (SENA, Centro Minero)"
    ws["A2"].alignment = alignment_center

    ws.merge_cells("A3:F3")
    ws["A3"] = "Proceso: Gestión de formación profesional integral"
    ws["A3"].alignment = alignment_center

    ws.merge_cells("A4:F4")
    ws["A4"] = "Procedimiento: Ejecución de la formación profesional integral"
    ws["A4"].alignment = alignment_center

    # Datos principales
    data_rows = [
        ["Programa de formación:", data.get("programa_formacion"), "Nº de ficha:", data.get("num_ficha"), "Fecha de inicio:", data.get("fecha_inicio_programa")],
        ["Nº Aprendices:", data.get("num_aprendices"), "Aprendiz vocero:", data.get("aprendiz"), "Instructor:", data.get("instructor")],
        ["Nº de contrato:", data.get("num_contrato"), "Fecha de entrega:", data.get("fecha"), "Dependencia:", data.get("dependencia")],
        ["Proyectos asociados:", data.get("proyecto"), "", "", "", ""]
    ]

    for row_index, row_data in enumerate(data_rows, start=5):
        for col_index, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.alignment = alignment_center if col_index % 2 == 0 else Alignment(horizontal="left")
            cell.font = bold_font if col_index % 2 != 0 else Font()
            cell.border = thin_border

    # Encabezados de tabla
    table_headers = ["Ítem", "Descripción del elemento", "Cantidad recibida", "Cantidad contratada", "Saldo pendiente de entrega", "Observaciones"]
    for col_index, header in enumerate(table_headers, start=1):
        cell = ws.cell(row=10, column=col_index, value=header)
        cell.font = bold_font
        cell.alignment = alignment_center
        cell.border = thin_border

    # Contenido de tabla
    table_data = data.get("items", [])
    for row_index, row_data in enumerate(table_data, start=11):
        for col_index, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.alignment = alignment_center
            cell.border = thin_border

    # Observaciones generales
    ws.merge_cells(f"A{len(table_data) + 12}:F{len(table_data) + 12}")
    ws[f"A{len(table_data) + 12}"] = f"Observaciones generales: {data.get('obs_general')}"
    ws[f"A{len(table_data) + 12}"].alignment = Alignment(horizontal="left")
    ws[f"A{len(table_data) + 12}"].font = bold_font

    # Guardar el archivo en memoria
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    # Enviar el archivo como respuesta
    return send_file(
        file_stream,
        as_attachment=True,
        download_name="Formato_SENA.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == '__main__':
    app.run(debug=True)
